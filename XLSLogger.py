from datetime import datetime
import os
import numpy as np
from openpyxl import Workbook, load_workbook

class XLSLogger:
   def __init__(self, file_path: str):
      self.file_path = file_path
      if not os.path.exists(self.file_path):
         self.workbook = Workbook()
         self.sheet = self.workbook.active
         self.sheet.title = 'Logs'
         self.row_index = 1
         self.workbook.save(self.file_path)
      else:
         self.workbook = load_workbook(self.file_path)
         self.sheet = self.workbook.active
         self.row_index = 1
         while self.sheet.cell(row=self.row_index, column=1).value is not None:
            self.row_index += 1

   def _normalize(self, val):
      if isinstance(val, (np.generic, np.ndarray)):
         return val.item()
      return val

   def _write_row(self, label: str, values: list):
      self.sheet.cell(row=self.row_index, column=1, value=label)
      for i, val in enumerate(values, start=2):
         self.sheet.cell(row=self.row_index, column=i, value=self._normalize(val))
      self.row_index += 1

   def writeFile(self, arr1, arr2):
      now = datetime.now().strftime("%d.%m.%Y %H:%M")
      self.sheet.cell(row=self.row_index, column=1, value=now)
      self.row_index += 1

      self._write_row("Control", arr1)
      self._write_row("Pure control", arr2)
      diffs = [abs(self._normalize(v1) - self._normalize(v2)) for v1, v2 in zip(arr1, arr2)]
      self._write_row("Loss", diffs)

      self.row_index += 1

      self.workbook.save(self.file_path)
